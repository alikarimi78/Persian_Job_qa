import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
import shutil

def split_xlsx_to_batches(input_file, batch_size=10, output_dir="batches"):
    """
    Split an xlsx file into batches of specified size, preserving headers.

    Args:
        input_file: Path to the xlsx file
        batch_size: Number of data rows per batch (default 10)
        output_dir: Output directory for batch files
    """
    input_path = Path(input_file)
    if not input_path.exists():
        raise FileNotFoundError(f"File not found: {input_file}")

    output_path = Path(output_dir)
    if output_path.exists():
        shutil.rmtree(output_path)
    output_path.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    headers = []
    for col in range(1, ws.max_column + 1):
        headers.append(ws.cell(row=1, column=col).value)

    total_data_rows = ws.max_row - 1
    batch_count = (total_data_rows + batch_size - 1) // batch_size

    print(f"Total data rows: {total_data_rows}")
    print(f"Batch size: {batch_size}")
    print(f"Total batches: {batch_count}")
    print(f"Output directory: {output_path.absolute()}")

    for batch_num in range(batch_count):
        start_row = 2 + (batch_num * batch_size)
        end_row = min(start_row + batch_size, ws.max_row + 1)

        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active

        for col, header in enumerate(headers, 1):
            new_ws.cell(row=1, column=col).value = header

        for i, row_idx in enumerate(range(start_row, end_row), 1):
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_idx, column=col).value
                new_ws.cell(row=i + 1, column=col).value = cell_value

        batch_filename = output_path / f"batch_{batch_num + 1:04d}.xlsx"
        new_wb.save(batch_filename)

        rows_in_batch = end_row - start_row
        print(f"  Batch {batch_num + 1}: {rows_in_batch} rows → {batch_filename.name}")

    print(f"\n✓ Successfully split into {batch_count} batches in {output_path}")

if __name__ == "__main__":
    split_xlsx_to_batches("data_extactor/onet_master_database_en.xlsx", batch_size=10, output_dir="data_extactor/batch_10")
