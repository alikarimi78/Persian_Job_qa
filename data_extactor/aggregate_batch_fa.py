from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

HERE = Path(__file__).resolve().parent

COLUMNS = ["job_code", "job_title", "aliases", "tools", "skills", "knowledge",
           "abilities", "work_context", "career_path_next", "description",
           "responsibilities"]

LIST_COLUMNS = ["aliases", "tools", "skills", "knowledge", "abilities",
                "work_context", "career_path_next", "responsibilities"]

PROSE_COLUMNS = ["job_title", "description"]

CANONICAL_COLUMNS = ["skills", "knowledge", "abilities", "work_context",
                     "career_path_next", "tools"]

MUST_BE_PERSIAN = ["job_title", "skills", "knowledge", "abilities", "work_context",
                   "career_path_next", "description", "responsibilities"]

PLACEHOLDERS = {"-", "--", "—", "–", ".", "‌", "نامشخص", "ندارد", "n/a", "N/A", "none"}

PERSIAN_RE = re.compile(r"[؀-ۿ]")


def items(value) -> list[str]:
    return [p.strip() for p in str(value).split("|") if p.strip()]


def blank(value) -> bool:
    v = str(value).strip()
    return not v or v.lower() in PLACEHOLDERS


def read_dir(directory: Path, suffix: str) -> tuple[pd.DataFrame, list[str]]:
    files = sorted(p for p in directory.glob("*.xlsx")
                   if not p.name.startswith("~$"))
    frames, names = [], []
    for path in files:
        name = path.stem[:-len(suffix)] if suffix and path.stem.endswith(suffix) else path.stem
        frame = pd.read_excel(path, dtype=str).fillna("")
        frame.columns = [str(c).strip() for c in frame.columns]
        frame["_batch"] = name
        frames.append(frame)
        names.append(name)
    if not frames:
        return pd.DataFrame(columns=COLUMNS + ["_batch"]), names
    return pd.concat(frames, ignore_index=True), names


def check(en: pd.DataFrame, fa: pd.DataFrame, en_names: list[str],
          fa_names: list[str]) -> tuple[list[str], dict]:
    hard: list[str] = []
    soft: dict = {}

    missing = [b for b in en_names if b not in set(fa_names)]
    extra = [b for b in fa_names if b not in set(en_names)]
    if missing:
        hard.append(f"{len(missing)} batches have no translation: {missing[:10]}"
                    + (" …" if len(missing) > 10 else ""))
    if extra:
        hard.append(f"{len(extra)} translated batches have no English original: {extra[:10]}")

    for label, frame in (("English", en), ("Persian", fa)):
        absent = [c for c in COLUMNS if c not in frame.columns]
        unknown = [c for c in frame.columns if c not in COLUMNS + ["_batch"]]
        if absent:
            hard.append(f"{label} side is missing columns: {absent}")
        if unknown:
            hard.append(f"{label} side has unexpected columns: {unknown}")
    if hard:
        return hard, soft

    if len(fa) != len(en):
        hard.append(f"row count: {len(fa)} translated against {len(en)} English")

    dup = fa.job_code[fa.job_code.duplicated()].tolist()
    if dup:
        hard.append(f"{len(dup)} duplicated job_code in the translation: {dup[:10]}")
    lost = [c for c in en.job_code if c not in set(fa.job_code)]
    added = [c for c in fa.job_code if c not in set(en.job_code)]
    if lost:
        hard.append(f"{len(lost)} job_code lost in translation: {lost[:10]}")
    if added:
        hard.append(f"{len(added)} job_code in the translation but not in the source: {added[:10]}")
    if not lost and not added and list(en.job_code) != list(fa.job_code):
        hard.append("job_code order differs between the two sides")

    empty = {}
    for column in COLUMNS:
        rows = fa.index[fa[column].map(blank)].tolist()
        if rows:
            empty[column] = [f"{fa.job_code.iloc[i]} ({fa._batch.iloc[i]})" for i in rows[:10]]
            hard.append(f"{column}: {len(rows)} empty or placeholder-only cells")
    soft["empty_cells"] = empty

    for column in PROSE_COLUMNS:
        rows = fa.index[fa[column].str.contains("|", regex=False)].tolist()
        if rows:
            hard.append(f"{column}: {len(rows)} cells contain '|', which is not a "
                        f"separator in a prose column: "
                        f"{[fa.job_code.iloc[i] for i in rows[:5]]}")

    if hard:
        return hard, soft

    left = en.set_index("job_code")
    right = fa.set_index("job_code")
    drift = {}
    for column in LIST_COLUMNS:
        rows = []
        for code in left.index:
            want, got = len(items(left[column][code])), len(items(right[column][code]))
            if want != got:
                rows.append({"job_code": code, "batch": right["_batch"][code],
                             "expected": want, "got": got})
        if rows:
            drift[column] = rows
    soft["item_drift"] = drift

    variants = {}
    for column in CANONICAL_COLUMNS:
        votes = {}
        for code in left.index:
            source, target = items(left[column][code]), items(right[column][code])
            if len(source) != len(target):
                continue
            for term, persian in zip(source, target):
                votes.setdefault(term, {}).setdefault(persian, 0)
                votes[term][persian] += 1
        split = {t: forms for t, forms in votes.items() if len(forms) > 1}
        if split:
            variants[column] = {
                "terms": len(split),
                "of": len(votes),
                "examples": [{"term": t, "forms": sorted(f, key=f.get, reverse=True)[:4]}
                             for t in sorted(split, key=lambda k: -sum(split[k].values()))[:5]],
            }
    soft["term_variants"] = variants

    untranslated = {}
    for column in COLUMNS[1:]:
        rows = [i for i in fa.index
                if str(fa[column].iloc[i]).strip() and not PERSIAN_RE.search(str(fa[column].iloc[i]))]
        if rows:
            untranslated[column] = {
                "rows": len(rows),
                "must_be_persian": column in MUST_BE_PERSIAN,
                "batches": sorted({fa._batch.iloc[i] for i in rows}),
                "examples": [fa.job_code.iloc[i] for i in rows[:5]],
            }
    soft["untranslated"] = untranslated
    return hard, soft


def write_xlsx(frame: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    frame.to_excel(path, index=False)

    book = load_workbook(path)
    sheet = book.active
    sheet.title = "onet_fa"
    fill = PatternFill("solid", start_color="4472C4", end_color="4472C4")
    head = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    body = Font(name="Arial", size=10)
    for cell in sheet[1]:
        cell.fill, cell.font = fill, head
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = body
            cell.alignment = Alignment(wrap_text=False)
    sheet.freeze_panes = "A2"
    for column in sheet.columns:
        width = max((len(str(c.value)) if c.value else 0) for c in column)
        sheet.column_dimensions[column[0].column_letter].width = min(width + 4, 60)
    book.save(path)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="in_dir", default=str(HERE / "batch_10_fa"),
                    help="the translated batches")
    ap.add_argument("--source", default=str(HERE / "batch_10"),
                    help="the English batches they were made from")
    ap.add_argument("--out", default=str(HERE / "onet_master_database_fa.xlsx"))
    ap.add_argument("--report", default="")
    ap.add_argument("--dry-run", action="store_true", help="check only, write nothing")
    ap.add_argument("--force", action="store_true",
                    help="write the merged file even if a hard check failed")
    args = ap.parse_args()

    in_dir, src_dir = Path(args.in_dir), Path(args.source)
    fa, fa_names = read_dir(in_dir, "_fa")
    en, en_names = read_dir(src_dir, "")
    if not en_names:
        print(f"error: no .xlsx in {src_dir}", file=sys.stderr)
        return 2
    if not fa_names:
        print(f"error: no .xlsx in {in_dir}", file=sys.stderr)
        return 2
    print(f"english : {len(en_names):4} batches, {len(en):5} rows  ({src_dir})")
    print(f"persian : {len(fa_names):4} batches, {len(fa):5} rows  ({in_dir})")

    hard, soft = check(en, fa, en_names, fa_names)

    print("\n--- hard checks ---")
    if hard:
        for line in hard:
            print(f"  FAIL  {line}")
    else:
        print("  all passed: batch set complete, row counts equal, job_code aligned, "
              "no empty cell, no '|' in prose")

    drift = soft.get("item_drift") or {}
    print("\n--- item counts against the English ---")
    if drift:
        for column, rows in sorted(drift.items()):
            lost = sum(r["expected"] - r["got"] for r in rows if r["expected"] > r["got"])
            print(f"  {column:18} {len(rows):4} cells differ (items lost {lost})"
                  f"  e.g. {rows[0]['job_code']} {rows[0]['expected']}→{rows[0]['got']}")
    else:
        print("  every cell has exactly the item count of its English original")

    variants = soft.get("term_variants") or {}
    print("\n--- one English term written more than one way ---")
    if variants:
        for column, info in sorted(variants.items()):
            print(f"  {column:18} {info['terms']:4} of {info['of']} terms have several forms")
            for e in info["examples"][:2]:
                print(f"       {e['term'][:40]:40} {' / '.join(e['forms'])}")
    else:
        print("  none: every English term has exactly one Persian form corpus-wide")

    untranslated = soft.get("untranslated") or {}
    print("\n--- cells with no Persian in them ---")
    if untranslated:
        for column, info in sorted(untranslated.items(),
                                   key=lambda kv: (not kv[1]["must_be_persian"], kv[0])):
            mark = "MUST BE PERSIAN" if info["must_be_persian"] else "allowed to be English"
            print(f"  {column:18} {info['rows']:4} rows  [{mark}]  "
                  f"{len(info['batches'])} batches, e.g. {info['examples'][:3]}")
    else:
        print("  none")

    if args.report:
        Path(args.report).write_text(
            json.dumps({"hard": hard, "soft": soft}, ensure_ascii=False, indent=2),
            encoding="utf-8")
        print(f"\nreport written to {args.report}")

    if hard and not args.force:
        print("\nnot writing the merged file: fix the hard failures above, or pass --force",
              file=sys.stderr)
        return 1
    if args.dry_run:
        print("\ndry run: nothing written")
        return 0

    merged = fa.sort_values("_batch", kind="stable").drop(columns=["_batch"])
    merged = merged[COLUMNS]
    write_xlsx(merged, Path(args.out))
    print(f"\nwrote {len(merged)} rows × {len(merged.columns)} columns to {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
