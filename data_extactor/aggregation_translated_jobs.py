import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import re

input_dir = Path("translated_xlsx")  # مسیر پوشه فایل‌ها

files = sorted(input_dir.glob("Translated_Summarized_Occupations_*.xlsx"))

dfs = []
for f in files:
    df = pd.read_excel(f)
    dfs.append(df)

merged_df = pd.concat(dfs, ignore_index=True)

def normalize_separators(val):
    """تبدیل جداکننده‌های مختلف به | استاندارد"""
    if not isinstance(val, str):
        return val
    # جایگزینی ، و , با |
    val = re.sub(r'\s*،\s*', ' | ', val)
    val = re.sub(r'\s*,\s*', ' | ', val)
    # پاک‌سازی | های تکراری یا اضافه
    val = re.sub(r'\s*\|\s*', ' | ', val)
    val = re.sub(r'(\|\s*)+', '| ', val)
    return val.strip(' |')

for col in merged_df.columns:
    merged_df[col] = merged_df[col].apply(normalize_separators)

output_path = "../Merged_Occupations.xlsx"
merged_df.to_excel(output_path, index=False)

wb = load_workbook(output_path)
ws = wb.active

header_fill = PatternFill("solid", start_color="4472C4", end_color="4472C4")
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
body_font = Font(name="Arial", size=10)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.font = body_font
        cell.alignment = Alignment(wrap_text=False)

ws.freeze_panes = "A2"

for col in ws.columns:
    max_len = max((len(str(c.value)) if c.value else 0) for c in col)
    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

wb.save(output_path)
print(f"Done: {len(merged_df)} records merged from {len(files)} files → {output_path}")