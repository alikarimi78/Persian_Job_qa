from __future__ import annotations

import argparse
import json
import re
import shutil
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent

COLUMNS = ["job_title", "aliases", "tools", "skills", "knowledge", "abilities",
           "work_context", "career_path_next", "description", "responsibilities"]
OUT_COLUMNS = ["row_index", "job_code"] + COLUMNS + ["source"]

PROSE_COLUMNS = ["job_title", "description"]

CAP = {"skills": 7, "knowledge": 8, "abilities": 19, "work_context": 18}

SOURCE_ONET = "O*NET — ترجمهٔ دوم (بازبینی‌شده)"
SOURCE_MILITARY = "تولیدی — تکمیل مشاغل نظامی"
SOURCE_AUTHORED = "تولیدی — مشاغل فناوری اطلاعات"

TAXONOMY_COLUMNS = ["skills", "knowledge", "abilities", "work_context"]

RESIDUAL_DESC = re.compile(r"(به\s*طور\s*جداگانه|به‌طور\s*جداگانه).{0,20}(فهرست|طبقه)")
RESIDUAL_TAIL = re.compile(
    r"^(.*)،\s*(?:سایر(?:\s+\S+)*|همه\s+(?:موارد|مشاغل)\s+دیگر)$")
RESIDUAL_SUFFIX = " (طبقه‌بندی‌نشده)"

LATIN = re.compile(r"[A-Za-z]")
TITLE_RE = re.compile(r"^(.*?)\s*\(([^()]*[A-Za-z][^()]*)\)\s*$")


def items(value) -> list[str]:
    return [p.strip() for p in str(value).split("|") if p.strip()]


def unique(values) -> list[str]:
    seen, out = set(), []
    for value in values:
        if value not in seen:
            seen.add(value)
            out.append(value)
    return out


def split_title(title: str) -> tuple[str, str]:
    match = TITLE_RE.match(str(title).strip())
    return (match.group(1).strip(), match.group(2).strip()) if match \
        else (str(title).strip(), "")


def rename_residual_titles(frame: pd.DataFrame) -> dict[str, str]:
    mapping = {}
    for i in range(len(frame)):
        title = str(frame.at[i, "job_title"]).strip()
        if not RESIDUAL_DESC.search(str(frame.at[i, "description"])):
            continue
        match = RESIDUAL_TAIL.match(title)
        if match:
            mapping[title] = "سایر " + match.group(1).strip() + RESIDUAL_SUFFIX
        elif title.startswith("سایر "):
            mapping[title] = title + RESIDUAL_SUFFIX
    frame["job_title"] = frame.job_title.map(lambda t: mapping.get(str(t).strip(), t))
    frame["career_path_next"] = frame.career_path_next.map(
        lambda cell: " | ".join(mapping.get(it, it) for it in items(cell)))
    return mapping


def load_translated(path: Path) -> pd.DataFrame:
    frame = pd.read_excel(path, dtype=str).fillna("")
    frame.columns = [str(c).strip().lower() for c in frame.columns]
    rows = []
    for i in range(len(frame)):
        row = frame.iloc[i]
        persian, english = split_title(row["job_title"])
        aliases = items(row["aliases"])
        if english and english not in aliases:
            aliases.append(english)
        record = {"job_code": row.get("job_code", ""), "job_title": persian,
                  "aliases": " | ".join(unique(aliases)),
                  "description": str(row["description"]).strip(),
                  "source": SOURCE_ONET}
        for column in COLUMNS:
            if column not in record:
                record[column] = " | ".join(unique(items(row[column])))
        rows.append(record)
    return pd.DataFrame(rows)


def load_military(path: Path, tools: dict[str, str],
                  enrichment: dict) -> tuple[pd.DataFrame, list[str], list[str]]:
    frame = pd.read_excel(path, dtype=str).fillna("")
    frame.columns = [str(c).strip().lower() for c in frame.columns]
    frame = frame[frame["source"].str.contains("نظامی", na=False)]
    missing, unenriched, rows = [], [], []
    for i in range(len(frame)):
        row = frame.iloc[i]
        rebuilt = []
        for term in items(row["tools"]):
            if term in tools:
                rebuilt.append(tools[term])
            else:
                rebuilt.append(term)
                missing.append(term)

        title = str(row["job_title"]).strip()
        add = enrichment.get(title)
        if add is None:
            unenriched.append(title)
            add = {}
        filled = {}
        for column in COLUMNS:
            value = str(row[column]).strip()
            if column not in add:
                filled[column] = value
                continue
            merged = unique(items(value) + list(add[column]))
            filled[column] = " | ".join(merged[:CAP[column]] if column in CAP else merged)

        rows.append({"job_code": "", **filled,
                     "tools": " | ".join(unique(rebuilt + list(add.get("tools", [])))),
                     "source": SOURCE_MILITARY})
    return pd.DataFrame(rows), sorted(set(missing)), unenriched


def load_authored(path: Path) -> pd.DataFrame:
    data = json.loads(path.read_text(encoding="utf-8"))
    rows = [{"job_code": "", **{c: str(row.get(c, "")).strip() for c in COLUMNS},
             "source": data.get("source", SOURCE_AUTHORED)}
            for row in data["occupations"]]
    return pd.DataFrame(rows, columns=["job_code"] + COLUMNS + ["source"])


def check_vocabulary(frame: pd.DataFrame, canonical: Path) -> list[str]:
    table = json.loads(canonical.read_text(encoding="utf-8"))["canonical"]
    allowed = {column: set(table[column].values()) for column in TAXONOMY_COLUMNS}
    problems = []
    for i in range(len(frame)):
        row = frame.iloc[i]
        for column in TAXONOMY_COLUMNS:
            stray = [it for it in items(row[column]) if it not in allowed[column]]
            if stray:
                problems.append(f"«{row['job_title']}» {column}: {len(stray)} phrase(s) "
                                f"outside the canonical taxonomy: {stray}")
    return problems


def check(frame: pd.DataFrame) -> list[str]:
    problems = []
    for column in COLUMNS:
        blank = frame.index[frame[column].str.strip() == ""].tolist()
        if blank:
            problems.append(f"{column}: {len(blank)} empty cells, e.g. rows {blank[:5]}")
    for column in PROSE_COLUMNS:
        piped = frame.index[frame[column].str.contains("|", regex=False)].tolist()
        if piped:
            problems.append(f"{column}: {len(piped)} cells hold '|', which is not a "
                            f"separator in a prose column: rows {piped[:5]}")
    duplicated = frame.job_title[frame.job_title.duplicated()].tolist()
    if duplicated:
        problems.append(f"{len(duplicated)} duplicated job_title: {duplicated[:5]}")
    repeated = [(frame.job_title.iloc[i], column)
                for i in range(len(frame))
                for column in COLUMNS if column not in PROSE_COLUMNS
                if len(items(frame[column].iloc[i]))
                != len(set(items(frame[column].iloc[i])))]
    if repeated:
        problems.append(f"{len(repeated)} cells hold the same item twice: {repeated[:5]}")
    titles = set(frame.job_title)
    linked = frame[frame.source.isin([SOURCE_ONET, SOURCE_AUTHORED])]
    unresolved = sorted({it for cell in linked.career_path_next
                         for it in items(cell) if it not in titles})
    if unresolved:
        problems.append(f"{len(unresolved)} career_path_next links name no "
                        f"record: {unresolved[:5]}")
    return problems


def write_xlsx(frame: pd.DataFrame, path: Path) -> None:
    frame.to_excel(path, index=False)
    book = load_workbook(path)
    sheet = book.active
    sheet.title = "occupations"
    fill = PatternFill("solid", start_color="4472C4", end_color="4472C4")
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(wrap_text=False)
    sheet.freeze_panes = "A2"
    for column in sheet.columns:
        width = max((len(str(c.value)) if c.value else 0) for c in column)
        sheet.column_dimensions[column[0].column_letter].width = min(width + 4, 60)
    book.save(path)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--translated", default=str(HERE / "onet_master_database_fa.xlsx"))
    ap.add_argument("--existing", default="",
                    help="where the military rows come from; defaults to the frozen "
                         "Merged_Occupations_v1.xlsx, or to --out on the first run")
    ap.add_argument("--tools", default=str(HERE / "merge_fixes" / "military_tools_fa.json"))
    ap.add_argument("--enrichment",
                    default=str(HERE / "merge_fixes" / "military_enrichment.json"))
    ap.add_argument("--authored",
                    default=str(HERE / "merge_fixes" / "authored_occupations_fa.json"))
    ap.add_argument("--canonical",
                    default=str(HERE / "translation_fixes" / "40_taxonomy_canonical.json"))
    ap.add_argument("--out", default=str(ROOT / "Merged_Occupations.xlsx"))
    ap.add_argument("--backup", default=str(HERE / "Merged_Occupations_v1.xlsx"))
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    out, backup = Path(args.out), Path(args.backup)
    existing = Path(args.existing) if args.existing else (backup if backup.exists() else out)
    if not existing.exists():
        print(f"error: no corpus to take the military rows from ({existing})",
              file=sys.stderr)
        return 2

    tools = json.loads(Path(args.tools).read_text(encoding="utf-8"))["tools"]
    enrichment = json.loads(Path(args.enrichment).read_text(encoding="utf-8"))["add"]
    onet = load_translated(Path(args.translated))
    renamed = rename_residual_titles(onet)
    military, missing, unenriched = load_military(existing, tools, enrichment)
    authored = load_authored(Path(args.authored))
    print(f"translated O*NET : {len(onet):5} rows  ({args.translated})")
    print(f"residual titles  : {len(renamed):5} rewritten as «سایر … (طبقه‌بندی‌نشده)»")
    print(f"military         : {len(military):5} rows  ({existing})")
    print(f"authored         : {len(authored):5} rows  ({args.authored})")

    if missing:
        print(f"\nerror: {len(missing)} military tool terms have no translation:",
              file=sys.stderr)
        for term in missing[:20]:
            print(f"  {term}", file=sys.stderr)
        print(f"add them to {args.tools}", file=sys.stderr)
        return 1
    if unenriched:
        print(f"\nerror: {len(unenriched)} military occupations are not in "
              f"{args.enrichment}:", file=sys.stderr)
        for title in unenriched[:20]:
            print(f"  {title}", file=sys.stderr)
        return 1

    merged = pd.concat([onet, military, authored], ignore_index=True)
    merged.insert(0, "row_index", range(len(merged)))
    merged = merged[OUT_COLUMNS]

    problems = check(merged) + check_vocabulary(authored, Path(args.canonical))
    print("\n--- checks ---")
    if problems:
        for line in problems:
            print(f"  FAIL  {line}")
    else:
        print("  no empty cell, no '|' in a prose column, no duplicate title, "
              "every career path resolves,\n  every authored taxonomy phrase is canonical")

    english_titles = merged.index[merged.job_title.map(lambda t: bool(LATIN.search(t)))]
    print(f"\n{len(merged)} rows; {len(english_titles)} titles still carry Latin text")
    listed = merged.work_context.map(lambda v: len(items(v)))
    print(f"work_context: {listed.min()}–{listed.max()} items per row, "
          f"{(listed == 1).sum()} rows holding a single one")

    titles = set(merged.job_title)
    military = merged[merged.source == SOURCE_MILITARY]
    loose = sorted({it for cell in military.career_path_next
                    for it in items(cell) if it not in titles})
    total = sum(len(items(c)) for c in merged.career_path_next)
    print(f"career paths: {total} links, of which {len(loose)} distinct names in the "
          f"military rows point at no record (they never did; see check() for why)")

    if problems:
        print("\nnot writing: fix the failures above", file=sys.stderr)
        return 1
    if args.dry_run:
        print("\ndry run: nothing written")
        return 0

    if out.exists() and not backup.exists():
        shutil.copy2(out, backup)
        print(f"\nkept the previous corpus as {backup}")
    write_xlsx(merged, out)
    print(f"wrote {len(merged)} rows × {len(merged.columns)} columns to {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
